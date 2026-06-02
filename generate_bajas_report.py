import os
import sys
import time
import django
import pymysql
import pandas as pd
import unicodedata
from datetime import datetime

# Setup django environment
sys.path.append("/home/edgar/ANAM/EjeCentral/eje_central_back")
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "eje_central_back.settings")
django.setup()

from plantilla.models import BajasSig

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

def clean_str(s):
    if not s:
        return ""
    s = str(s).strip()
    s = "".join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    return " ".join(s.upper().split())

def clean_date(d):
    if not d:
        return ""
    d = str(d).strip()
    if len(d) >= 10:
        return d[:10]  # Obtener YYYY-MM-DD
    return d

def format_diff(field_name, val_ext, val_loc):
    val_ext_str = str(val_ext or '').strip()
    val_loc_str = str(val_loc or '').strip()
    return f"Diferencia en {field_name}: '{val_ext_str}' en Externo (Zafiro MOV_TOTAL) vs '{val_loc_str}' en Local (EjeCentral BAJAS_SIG)"

def main():
    print("=== INICIANDO EXTRACCIÓN Y COMPARACIÓN DE DATOS CON LÓGICA DE MOVIMIENTOS ===")
    
    # 1. Fetch local data
    print("Cargando registros locales de BAJAS_SIG...")
    bajas_loc = list(BajasSig.objects.values(
        "no_empleado", "posicion", "nombre_completo", 
        "primer_apellido", "segundo_apellido", "motivo", "motivo_descr",
        "fecha_efectiva", "fecha_aplicacion", "ultima_actualizacion"
    ))
    df_loc = pd.DataFrame(bajas_loc)
    print(f"  -> {len(df_loc)} registros locales cargados.")
    
    # Normalizar campos de búsqueda local
    df_loc["emp_id_norm"] = df_loc["no_empleado"].apply(lambda x: clean_str(x).lstrip('0'))
    df_loc["pos_norm"] = df_loc["posicion"].apply(lambda x: clean_str(x).lstrip('0'))
    df_loc["paterno_norm"] = df_loc["primer_apellido"].apply(clean_str)
    df_loc["materno_norm"] = df_loc["segundo_apellido"].apply(clean_str)
    df_loc["fullname_norm"] = df_loc["nombre_completo"].apply(clean_str)
    df_loc["motivo_code"] = df_loc["motivo"].apply(clean_str)
    df_loc["motivo_name"] = df_loc["motivo_descr"].apply(clean_str)
    df_loc["date_efva_norm"] = df_loc["fecha_efectiva"].apply(clean_date)
    df_loc["date_cap_norm"] = df_loc["fecha_aplicacion"].apply(clean_date)
    
    # Agrupar por empleado local para búsquedas rápidas
    bajas_by_emp = {}
    for _, row in df_loc.iterrows():
        emp_id = row["emp_id_norm"]
        if emp_id not in bajas_by_emp:
            bajas_by_emp[emp_id] = []
        bajas_by_emp[emp_id].append(row.to_dict())
        
    # 2. Fetch external data
    print("Cargando registros externos de MOV_TOTAL (Acción = baja)...")
    conn = pymysql.connect(
        host="168.231.73.222",
        port=3306,
        user="omar.ramirez",
        password="Raal1011.",
        database="safirho_db",
        charset="utf8mb4"
    )
    with conn.cursor(pymysql.cursors.DictCursor) as cursor:
        cursor.execute("SELECT * FROM MOV_TOTAL WHERE `Acción (Nombre)` = 'baja'")
        rows_ext = cursor.fetchall()
    conn.close()
    
    df_ext = pd.DataFrame(rows_ext)
    print(f"  -> {len(df_ext)} registros externos cargados.")
    
    # 3. Comparación registro por registro
    print("Comparando registros...")
    coincidencias_totales = []
    diferencias = []
    sin_bajas = []
    
    for _, r_ext in df_ext.iterrows():
        emp_id_raw = r_ext.get("Id_empl") or ""
        emp_id_norm = clean_str(emp_id_raw).lstrip('0')
        
        pos_raw = r_ext.get("Posición") or ""
        pos_norm = clean_str(pos_raw).lstrip('0')
        
        ext_nombre = clean_str(r_ext.get("Nombre"))
        ext_paterno = clean_str(r_ext.get("Paterno"))
        ext_materno = clean_str(r_ext.get("Apellido Matern"))
        
        ext_motivo_code = clean_str(r_ext.get("Motivo"))
        ext_motivo_name = clean_str(r_ext.get("Motivo (Nombre)"))
        
        f_efva_raw = r_ext.get("F/Efva") or ""
        f_captura_raw = r_ext.get("F/Captura") or ""
        
        ext_date_efva = clean_date(f_efva_raw)
        ext_date_cap = clean_date(f_captura_raw)
        
        # Estructura del registro para el reporte
        reg_base = {
            "Id Empleado (Ext)": emp_id_raw,
            "Posición (Ext)": pos_raw,
            "Nombre Completo (Ext)": f"{r_ext.get('Nombre') or ''} {r_ext.get('Paterno') or ''} {r_ext.get('Apellido Matern') or ''}".strip(),
            "Motivo (Ext)": f"{ext_motivo_code} - {r_ext.get('Motivo (Nombre)') or ''}",
            "Fecha Efectiva (Ext)": f_efva_raw,
            "Fecha Captura (Ext)": f_captura_raw,
        }
        
        # Caso A: No existe el empleado en absoluto
        if emp_id_norm not in bajas_by_emp:
            sin_bajas.append(reg_base)
            continue
            
        # Caso B: Existe el empleado, buscamos si hay una coincidencia de movimiento
        # Criterio: Mismo Motivo AND Mismas Fechas (Efectiva y Captura)
        candidates = bajas_by_emp[emp_id_norm]
        movimiento_encontrado = None
        
        for cand in candidates:
            # Comprobar Motivo
            mismo_motivo = (ext_motivo_code == cand["motivo_code"] or ext_motivo_name == cand["motivo_name"])
            
            # Comprobar Fecha Efectiva
            cand_date_efva = cand["date_efva_norm"]
            misma_fecha_efva = (ext_date_efva == cand_date_efva and ext_date_efva != "")
            
            # Comprobar Fecha Captura
            cand_date_cap = cand["date_cap_norm"] or clean_date(cand["ultima_actualizacion"])
            misma_fecha_cap = (ext_date_cap == cand_date_cap and ext_date_cap != "")
            
            if mismo_motivo and misma_fecha_efva and misma_fecha_cap:
                movimiento_encontrado = cand
                break
                
        if movimiento_encontrado is None:
            # Si difiere en Motivo o Fechas, es un movimiento distinto -> Sin Registro Local
            sin_bajas.append(reg_base)
        else:
            # Si es el mismo movimiento, verificamos posición y nombre para determinar si es Coincidencia Total o Diferencia
            cand = movimiento_encontrado
            mismo_posicion = (pos_norm == cand["pos_norm"])
            mismo_nombre = (ext_paterno == cand["paterno_norm"] and ext_materno == cand["materno_norm"] and ext_nombre in cand["fullname_norm"])
            
            reg_match = {
                **reg_base,
                "No Empleado (Loc)": cand["no_empleado"],
                "Posición (Loc)": cand["posicion"],
                "Nombre Completo (Loc)": cand["nombre_completo"],
                "Motivo (Loc)": f"{cand['motivo']} - {cand['motivo_descr']}",
                "Fecha Efectiva (Loc)": cand["fecha_efectiva"],
                "Fecha Captura (Loc)": cand["fecha_aplicacion"] or clean_date(cand["ultima_actualizacion"]),
            }
            
            if mismo_posicion and mismo_nombre:
                coincidencias_totales.append(reg_match)
            else:
                mismatches = []
                if not mismo_posicion:
                    mismatches.append(format_diff("Posición", pos_raw, cand["posicion"]))
                if not mismo_nombre:
                    if ext_paterno != cand["paterno_norm"]:
                        mismatches.append(format_diff("Paterno", r_ext.get('Paterno'), cand["primer_apellido"]))
                    if ext_materno != cand["materno_norm"]:
                        mismatches.append(format_diff("Materno", r_ext.get('Apellido Matern'), cand["segundo_apellido"]))
                    if ext_nombre not in cand["fullname_norm"]:
                        mismatches.append(format_diff("Nombre de Pila", r_ext.get('Nombre'), cand["nombre_completo"]))
                
                reg_match["Diferencias Encontradas"] = " | ".join(mismatches)
                diferencias.append(reg_match)
            
    print(f"Resultados finales de la comparación:")
    print(f"  - Coincidencias Totales (100% igual): {len(coincidencias_totales)}")
    print(f"  - Diferencias (Mismo Movimiento pero difiere Posición o Nombre): {len(diferencias)}")
    print(f"  - Sin Registro Local (Diferente Movimiento o No Existe): {len(sin_bajas)}")
    
    # 4. Generación del reporte en Excel
    print("Creando archivo Excel con openpyxl...")
    wb = Workbook()
    
    # Paleta de colores profesionales
    COLOR_HEADER_BG = "1F4E78"     # Azul Acero
    COLOR_HEADER_FG = "FFFFFF"     # Blanco
    COLOR_ZEBRA_BG = "F2F6F9"      # Gris Azulado Claro
    COLOR_DIF_BG = "FFEBEE"        # Rojo Claro para resaltar diferencias
    COLOR_DIF_FG = "C62828"        # Rojo Oscuro
    
    font_family = "Segoe UI"
    
    # --- HOJA 1: RESUMEN ---
    ws_summary = wb.active
    ws_summary.title = "Resumen"
    ws_summary.views.sheetView[0].showGridLines = True
    
    # Título principal
    ws_summary["A1"] = "REPORTE DE COMPARACIÓN DE BAJAS"
    ws_summary["A1"].font = Font(name=font_family, size=16, bold=True, color="1F4E78")
    ws_summary["A2"] = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws_summary["A2"].font = Font(name=font_family, size=10, italic=True, color="555555")
    
    # Tabla de resumen
    ws_summary["A4"] = "Estatus de Comparación"
    ws_summary["B4"] = "Cantidad"
    ws_summary["C4"] = "Porcentaje"
    
    for col in ["A4", "B4", "C4"]:
        ws_summary[col].font = Font(name=font_family, size=11, bold=True, color=COLOR_HEADER_FG)
        ws_summary[col].fill = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
        ws_summary[col].alignment = Alignment(horizontal="center")
        
    data_resumen = [
        ("Coincidencia Total (100% igual)", len(coincidencias_totales)),
        ("Diferencias Encontradas (Mismo Movimiento, difiere Posición/Nombre)", len(diferencias)),
        ("Sin Registro Local (Diferente Movimiento o No Existe)", len(sin_bajas)),
    ]
    
    for i, (cat, val) in enumerate(data_resumen, start=5):
        ws_summary[f"A{i}"] = cat
        ws_summary[f"B{i}"] = val
        ws_summary[f"C{i}"] = f"=B{i}/$B$8"
        
        # Estilos filas de datos
        ws_summary[f"A{i}"].font = Font(name=font_family, size=10)
        ws_summary[f"B{i}"].font = Font(name=font_family, size=10)
        ws_summary[f"C{i}"].font = Font(name=font_family, size=10)
        
        ws_summary[f"B{i}"].alignment = Alignment(horizontal="right")
        ws_summary[f"C{i}"].alignment = Alignment(horizontal="right")
        ws_summary[f"B{i}"].number_format = "#,##0"
        ws_summary[f"C{i}"].number_format = "0.0%"
        
        # Borde inferior fino
        thin_border = Border(bottom=Side(style='thin', color='DDDDDD'))
        for col in ["A", "B", "C"]:
            ws_summary[f"{col}{i}"].border = thin_border
            
    # Totalizadores
    ws_summary["A8"] = "Total Analizado (MOV_TOTAL)"
    ws_summary["B8"] = "=SUM(B5:B7)"
    ws_summary["C8"] = "=SUM(C5:C7)"
    
    for col in ["A8", "B8", "C8"]:
        ws_summary[col].font = Font(name=font_family, size=11, bold=True)
        ws_summary[col].border = Border(top=Side(style='thin', color='000000'), bottom=Side(style='double', color='000000'))
    ws_summary["B8"].number_format = "#,##0"
    ws_summary["C8"].number_format = "0.0%"
    ws_summary["B8"].alignment = Alignment(horizontal="right")
    ws_summary["C8"].alignment = Alignment(horizontal="right")
    
    # Agregar Gráfico de Barras
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Estatus de Coincidencia de Bajas (Clasificación por Movimientos)"
    chart.y_axis.title = "Registros"
    chart.x_axis.title = "Estatus"
    chart.width = 16
    chart.height = 10
    
    data_ref = Reference(ws_summary, min_col=2, min_row=4, max_row=7)
    cats_ref = Reference(ws_summary, min_col=1, min_row=5, max_row=7)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.legend = None
    
    ws_summary.add_chart(chart, "E4")
    
    # Autoajuste de columnas en Resumen
    for col in ws_summary.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_summary.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    # --- HELPER PARA CREAR HOJAS DE DETALLE ---
    def create_detail_sheet(name, data_list, has_loc=True, has_dif=False):
        ws = wb.create_sheet(title=name)
        ws.views.sheetView[0].showGridLines = True
        ws.freeze_panes = "A2"
        
        if not data_list:
            ws["A2"] = "No se encontraron registros para esta categoría."
            ws["A2"].font = Font(name=font_family, size=11, italic=True)
            return
            
        # Obtener columnas
        headers = list(data_list[0].keys())
        
        # Escribir cabeceras
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = Font(name=font_family, size=11, bold=True, color=COLOR_HEADER_FG)
            cell.fill = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
        ws.row_dimensions[1].height = 28
        
        # Escribir datos
        thin_border = Border(
            left=Side(style='thin', color='E5E5E5'),
            right=Side(style='thin', color='E5E5E5'),
            top=Side(style='thin', color='E5E5E5'),
            bottom=Side(style='thin', color='E5E5E5')
        )
        
        for row_idx, data in enumerate(data_list, start=2):
            ws.row_dimensions[row_idx].height = 20
            is_zebra = (row_idx % 2 == 0)
            row_fill = PatternFill(start_color=COLOR_ZEBRA_BG, end_color=COLOR_ZEBRA_BG, fill_type="solid") if is_zebra else None
            
            for col_idx, h in enumerate(headers, start=1):
                val = data[h]
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = Font(name=font_family, size=10)
                cell.border = thin_border
                if row_fill:
                    cell.fill = row_fill
                
                # Formato y alineación de columnas específicas
                if "Id Empleado" in h or "No Empleado" in h or "Posición" in h:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.number_format = "@"
                elif "Fecha" in h:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    
                # Si es la columna de diferencias, darle un formato especial resaltado
                if h == "Diferencias Encontradas":
                    cell.fill = PatternFill(start_color=COLOR_DIF_BG, end_color=COLOR_DIF_BG, fill_type="solid")
                    cell.font = Font(name=font_family, size=10, bold=True, color=COLOR_DIF_FG)
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    
        # Auto-ajuste de anchos de columna
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            if col[0].value == "Diferencias Encontradas":
                ws.column_dimensions[col_letter].width = 60
            else:
                max_len = max(len(str(cell.value or '')) for cell in col)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
                
    # Crear las hojas de detalle
    print("Generando pestaña de Coincidencias Totales...")
    create_detail_sheet("Coincidencias Totales", coincidencias_totales)
    
    print("Generando pestaña de Diferencias...")
    create_detail_sheet("Diferencias", diferencias, has_dif=True)
    
    print("Generando pestaña de Sin Registro Local...")
    create_detail_sheet("Sin Registro Local", sin_bajas, has_loc=False)
    
    # 5. Guardar libro
    out_file = "/home/edgar/ANAM/EjeCentral/eje_central_back/Reporte_Comparacion_Bajas.xlsx"
    print(f"Guardando archivo final en: {out_file}")
    wb.save(out_file)
    print("¡Reporte con nueva lógica generado con éxito!")

if __name__ == "__main__":
    main()
